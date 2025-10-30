# main.py
import os
import threading
import time as t
from datetime import datetime, timedelta, time as dt_time, date as dt_date

import pytz
import telebot
from dotenv import load_dotenv
from openpyxl import load_workbook

# ---------------------------
# Load bot token
# ---------------------------
load_dotenv()
BOT_TOKEN = os.getenv("BOT_TOKEN")
if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN not found in .env")

bot = telebot.TeleBot(BOT_TOKEN, parse_mode="Markdown")

# ---------------------------
# Target chat / topic
# ---------------------------
# CHAT_ID = -1002257320998
# TOPIC_ID = 751
CHAT_ID = -1002635519712
TOPIC_ID = 434

# ---------------------------
# Timezone (Singapore)
# ---------------------------
SGT = pytz.timezone("Asia/Singapore")

# ---------------------------
# Configurable file path
# ---------------------------
SCHEDULE_FILE = "schedule.xlsx"

# ---------------------------
# Utility: check if weekend (Saturday/Sunday)
# ---------------------------
def is_weekend(date_obj: dt_date) -> bool:
    return date_obj.weekday() >= 5  # Saturday=5, Sunday=6

# ---------------------------
# Read schedule.xlsx for a specific date
# ---------------------------
def get_schedule_for_date(target_date: dt_date, file_path: str = SCHEDULE_FILE):
    try:
        wb = load_workbook(filename=file_path, data_only=True)
        sheet = wb.active
    except Exception as e:
        print(f"Failed to open '{file_path}': {e}")
        return None, None, None

    schedule_data = {"Morning": [], "Afternoon": [], "Night": []}
    date_str = None
    day_str = None

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 7:
            continue
        date_val, day, level, shift, student_id, name, building = row[:7]

        # normalize date_val
        try:
            if isinstance(date_val, datetime):
                row_date = date_val.date()
            elif isinstance(date_val, dt_date):
                row_date = date_val
            else:
                s = str(date_val).strip()
                parsed = None
                for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
                    try:
                        parsed = datetime.strptime(s, fmt).date()
                        break
                    except Exception:
                        pass
                if parsed is None:
                    continue
                row_date = parsed
        except Exception:
            continue

        if row_date == target_date:
            date_str = row_date.strftime("%Y/%m/%d")
            day_str = str(day).strip() if day else row_date.strftime("%A")
            if name and str(name).strip() and str(name).strip().upper() != "NIL":
                shift_name = str(shift).strip().capitalize() if shift else "Afternoon"
                level_label = str(level).strip() if level else ""
                schedule_data.setdefault(shift_name, []).append(f"{str(name).strip()} ({level_label})")

    if not any(schedule_data.values()):
        return None, None, None

    return date_str, day_str, schedule_data

# ---------------------------
# Format schedule message
# ---------------------------
def format_message(date_str: str, day_str: str, schedule_data: dict, label: str) -> str:
    lines = []
    lines.append(f"*Schedule for {label}*")
    lines.append(f"Date: {date_str}")
    lines.append(f"Day: {day_str}")
    lines.append("")  # blank line

    for shift in ["Morning", "Afternoon", "Night"]:
        lines.append(f"*{shift} shift*")
        names = schedule_data.get(shift, [])
        if names:
            lines.extend(names)
        else:
            lines.append("_No one scheduled_")
        lines.append("")  # blank line

    return "\n".join(lines)

# ---------------------------
# Send schedule message
# ---------------------------
def send_schedule(target_date: dt_date, label: str, allow_send_on_weekends_for_manual: bool = True):
    weekday = target_date.weekday()

    # Saturday or Sunday → CLOSED message for auto-send
    if weekday >= 5:
        msg = f"📅 *Schedule for {label}*\n"
        msg += f"Date: {target_date.strftime('%Y/%m/%d')}\n"
        msg += f"Day: {target_date.strftime('%A')}\n\n"
        msg += "*ProjectHub CLOSED*\n\n"
        msg += f"No schedule tomorrow as it is {target_date.strftime('%A')}. Thank you all for your service, enjoy the weekend everyone!"
        try:
            bot.send_message(chat_id=CHAT_ID, text=msg, message_thread_id=TOPIC_ID)
            print(f"Closed day message sent for {label} ({target_date.strftime('%A')})")
        except Exception as e:
            print(f"Failed to send closed message: {e}")
        return

    # Normal schedule
    date_str, day_str, schedule_data = get_schedule_for_date(target_date)
    if not date_str:
        print(f"No schedule found for {label} ({target_date.isoformat()}).")
        return

    msg = format_message(date_str, day_str, schedule_data, label)
    try:
        bot.send_message(chat_id=CHAT_ID, text=msg, message_thread_id=TOPIC_ID)
        print(f"{label.capitalize()} schedule sent to Chat {CHAT_ID} (Topic {TOPIC_ID}) for {date_str}.")
    except Exception as e:
        print(f"Failed to send schedule message for {label}: {e}")

# ---------------------------
# Scheduler thread for auto-send
# ---------------------------
def scheduler_thread():
    print("Scheduler thread started. Auto-send at 17:50 SGT (Sat/Sun = CLOSED, Mon–Fri = normal).")
    while True:
        now = datetime.now(SGT)
        today_date = now.date()
        target_dt = SGT.localize(datetime.combine(today_date, dt_time(17, 50, 0)))

        if now >= target_dt:
            target_dt += timedelta(days=1)

        wait_seconds = (target_dt - now).total_seconds()
        print(f"Next auto-send scheduled at: {target_dt.strftime('%Y-%m-%d %H:%M:%S %Z')}")
        t.sleep(wait_seconds + 0.5)

        tomorrow_date = (datetime.now(SGT) + timedelta(days=1)).date()
        send_schedule(tomorrow_date, "tomorrow", allow_send_on_weekends_for_manual=True)

# ---------------------------
# Telegram command handlers
# ---------------------------
@bot.message_handler(commands=["today_schedule"])
def handle_today_schedule(message):
    try:
        today = datetime.now(SGT).date()
        send_schedule(today, "today", allow_send_on_weekends_for_manual=True)
    except Exception as e:
        print(f"Exception in today_schedule handler: {e}")

@bot.message_handler(commands=["tomorrow_schedule"])
def handle_tomorrow_schedule(message):
    try:
        tomorrow = (datetime.now(SGT) + timedelta(days=1)).date()
        send_schedule(tomorrow, "tomorrow", allow_send_on_weekends_for_manual=True)
    except Exception as e:
        print(f"Exception in tomorrow_schedule handler: {e}")

# ---------------------------
# Main
# ---------------------------
if __name__ == "__main__":
    t_thread = threading.Thread(target=scheduler_thread, daemon=True)
    t_thread.start()

    print("Bot is running.")
    print("Commands: /today_schedule  /tomorrow_schedule")
    print("Auto-send: tomorrow's schedule at 17:50 SGT (Sat/Sun = CLOSED, Mon–Fri = normal)")

    bot.infinity_polling()
